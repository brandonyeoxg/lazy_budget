import argparse
import logging

from datetime import datetime
from pathlib import Path
from monopoly.pdf import PdfDocument, PdfParser
from monopoly.banks import BankDetector, banks
from monopoly.generic import GenericBank
from monopoly.pipeline import Pipeline
from openpyxl import Workbook, load_workbook

logging.basicConfig(
    level=logging.DEBUG,
)
logger = logging.getLogger(__name__)


class Sheet:
    def __init__(self, title, columns, transactions):
        self.title = title
        self.columns = columns
        self.transactions = transactions


def get_statement_from_file(file_path: Path):
    document = PdfDocument(file_path=file_path)
    document.unlock_document()
    detector = BankDetector(document)
    bank = detector.detect_bank(banks) or GenericBank
    pdf_parser = PdfParser(bank, document)
    pipeline = Pipeline(pdf_parser)

    statement = pipeline.extract(safety_check=False)
    transactions = pipeline.transform(statement)
    transactions_as_dict = [transaction.as_raw_dict() for transaction in transactions]
    return statement


def save_as_report(name, statements):
    wb = Workbook()
    wb.remove(wb.active)
    for statement in sorted(statements, key=lambda s: s.statement_date):
        title = statement.statement_date.strftime("%b_%Y")
        ws = wb.create_sheet(title)
        ws.append(statement.columns)
        COLUMNS = [str(column) for column in statement.columns]
        for transaction in statement.transactions:
            d = transaction.as_raw_dict()
            ws.append([d.get(c, "") for c in COLUMNS])
    wb.save(name)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--report_name", help="name of the report")
    parser.add_argument("--files", nargs="*", help="bank statement in pdf files")
    parser.add_argument("--folder", help="folder to process the entire statements")
    args = parser.parse_args()
    if args.files and args.folder:
        logger.error("Both flags cannot be on at the same time")
        return

    name = datetime.now().strftime("report_%d-%m-%Y_%H-%M-%S.xlsx")
    if args.report_name:
        name = args.report_name
    statements = list()

    if args.files:
        for file in args.files:
            statements.append(get_statement_from_file(file))
        save_as_report(name, statements)
        return

    if args.folder:
        folder = Path(args.folder)
        files = [f for f in folder.iterdir() if f.is_file()]
        for file in files:
            statements.append(get_statement_from_file(file))
        save_as_report(name, statements)
        return


if __name__ == "__main__":
    main()
